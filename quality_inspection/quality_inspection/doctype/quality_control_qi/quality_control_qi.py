# Copyright (c) 2024, GreyCube Technologies and contributors
# For license information, please see license.txt

import frappe
from frappe import _
from frappe.model.document import Document
from frappe.model.mapper import get_mapped_doc
from frappe.utils import cstr, flt
import openpyxl
from openpyxl.styles import Font, Alignment, PatternFill
from openpyxl.styles import Border, Side
import os

PASS_STATUS = ["Pass"]
UNDETERMINED = ["Undetermined"]
TODO_STATUS = ["To Do"]
class QualityControlQI(Document):

	def onload(self):
		self.get_pallet_information_html()

	def validate(self):
		self.set_pallet_and_moisture_default_value()
		self.set_attachments_details()
		self.set_color_count_for_color_match_and_embossing()
		# self.validate_over_wax_and_edge_paint_child_table()
		
	def get_pallet_information_html(self):

		template_path = "templates/pallet_information.html"

		pallet_html = frappe.render_template(template_path,  
									   dict(total_attach=self.pallet_total_attachment, pending=self.pallet_pending_attachment, 
				 							pass_ration=self.pallet_pass_ration,color=None, undetermined_ratio=self.pallet_undetermined_ratio, todo_ration=self.pallet_todo_ratio))  
		inner_outer_html = frappe.render_template(template_path,  
										dict(total_attach=self.inner_total_attachment, pending=self.inner_pending_attachment, 
			   								pass_ration=self.inner_pass_ration,color=None, undetermined_ratio=self.inner_undetermined_ratio, todo_ration=self.inner_todo_ratio))
		color_match_html = frappe.render_template(template_path,
										dict(total_attach=self.color_total_attachment, pending=self.color_pending_attachment, 
			   								pass_ration=self.color_pass_ration,color=self.color_count, undetermined_ratio=self.color_undetermined_ratio, todo_ration=self.color_todo_ratio))
		over_wax_html = frappe.render_template(template_path,  
										dict(total_attach=self.over_total_attachment, pending=self.over_pending_attachment, 
			   								pass_ration=self.over_pass_ration,color=None, undetermined_ratio=self.over_undetermined_ratio, todo_ration=self.over_todo_ratio))
		gloss_level_html = frappe.render_template(template_path,  
										dict(total_attach=self.gloss_total_attachment, pending=self.gloss_pending_attachment, 
			   								pass_ration=self.gloss_pass_ration,color=None, undetermined_ratio=self.gloss_undetermined_ratio, todo_ration=self.gloss_todo_ratio))
		moisture_html = frappe.render_template(template_path,  
										dict(total_attach=self.moisture_total_attachment, pending=self.moisture_pending_attachment, 
			   								pass_ration=self.moisture_pass_ration,color=None, undetermined_ratio=self.moisture_undetermined_ratio, todo_ration=self.moisture_todo_ratio))
		open_box_html = frappe.render_template(template_path,  
										dict(total_attach=self.open_total_attachment, pending=self.open_pending_attachment, 
			   								pass_ration=self.open_pass_ration,color=None, undetermined_ratio=self.open_undetermined_ratio, todo_ration=self.open_todo_ratio))
		width_thickness_html = frappe.render_template(template_path,  
										dict(total_attach=self.width_total_attachment, pending=self.width_pending_attachment, 
			   								pass_ration=self.width_pass_ration,color=None, undetermined_ratio=self.open_undetermined_ratio, todo_ration=self.width_todo_ratio))

		self.set_onload("pallet_html", pallet_html)
		self.set_onload("inner_outer_html", inner_outer_html)
		self.set_onload("color_match_html", color_match_html)
		self.set_onload("over_wax_html", over_wax_html)
		self.set_onload("gloss_level_html", gloss_level_html)
		self.set_onload("moisture_html", moisture_html)
		self.set_onload("open_box_html", open_box_html)
		self.set_onload("width_thickness_html", width_thickness_html)

	def calculate_attachments_details(self, child_table, attach_field_list, select_field_list):
		
		total_attachments = 0
		pendings = 0
		total_pass = 0
		pass_ration = 0
		total_undetermined = 0
		undetermined_ratio = 0
		total_todo = 0
		todo_ration = 0

		if self.get("no_of_po") and self.get("no_of_po") > 0:
			if child_table == "pallet_details":
				total_select_fields = len(self.pallet_details)	

				for row in self.get(child_table):

					for attach in attach_field_list:
						if not row.get(attach):
							pendings = pendings + 1
						total_attachments = total_attachments + 1

					for select in select_field_list:
						if row.get(select) in PASS_STATUS:
							total_pass = total_pass + 1
						elif row.get(select) in UNDETERMINED:
							total_undetermined = total_undetermined + 1
						elif row.get(select) in TODO_STATUS:
							total_todo = total_todo + 1
						else:
							pass

				pass_ration = ((total_pass * 100)/ total_select_fields)
				undetermined_ratio = ((total_undetermined * 100) / total_select_fields)
				todo_ration = ((total_todo * 100) / total_select_fields)

			else:
				table_length = 0
				for i in range(self.no_of_po):
					child_table_name=child_table+cstr(i+1)
					
					if len(self.get(child_table_name)) > 0:
						table_length = table_length +  len(self.get(child_table_name))
						for row in self.get(child_table_name):

							for attach in attach_field_list:
								if not row.get(attach):
									pendings = pendings + 1
								total_attachments = total_attachments + 1

							for select in select_field_list:
								if row.get(select) in PASS_STATUS:
									total_pass = total_pass + 1
								elif row.get(select) in UNDETERMINED:
									total_undetermined = total_undetermined + 1
								elif row.get(select) in TODO_STATUS:
									total_todo = total_todo + 1
								else:
									pass


				print(total_todo, "========total_todo")
				total_select_fields = table_length * len(select_field_list)
				print(total_select_fields, "=========total_select_fields")

				pass_ration = ((total_pass * 100)/ total_select_fields)
				undetermined_ratio = ((total_undetermined * 100) / total_select_fields)
				todo_ration = ((total_todo * 100) / total_select_fields)
		
		return total_attachments, pendings, pass_ration, undetermined_ratio, todo_ration
	
	def set_attachments_details(self):
		pallet = self.calculate_attachments_details("pallet_details", ['width', 'height'], ['button_select'])
		self.pallet_total_attachment = pallet[0]
		self.pallet_pending_attachment = pallet[1]
		self.pallet_pass_ration = flt((pallet[2]), 2)
		self.pallet_undetermined_ratio = flt((pallet[3]), 2)
		self.pallet_todo_ratio = flt((pallet[4]), 2)

		inner = self.calculate_attachments_details("inner_and_outer_carton_details_", ['end_label'],
											 ['hologram_select', 'carb_select', 'floor_select', 'shink_wrap_select', 'insert_sheet_select'])
		self.inner_total_attachment = inner[0]
		self.inner_pending_attachment = inner[1]
		self.inner_pass_ration = flt((inner[2]), 2)
		self.inner_undetermined_ratio = flt((inner[3]), 2)
		self.inner_todo_ratio = flt((inner[4]), 2)

		color = self.calculate_attachments_details("color_match_and_embossing_details_", ['master_sample', 'finished_board'], ['results_select', 'embossing_select'])
		self.color_total_attachment = color[0]
		self.color_pending_attachment = color[1]
		self.color_pass_ration = flt((color[2]), 2)
		self.color_undetermined_ratio = flt((color[3]), 2)
		self.color_todo_ratio = flt((color[4]), 2)

		over = self.calculate_attachments_details("over_wax_and_edge_paint_", ['finished_board'], ['over_wax_select', 'edge_paint_select'])
		self.over_total_attachment = over[0]
		self.over_pending_attachment = over[1]
		self.over_pass_ration = flt((over[2]), 2)
		self.over_undetermined_ratio = flt((over[3]), 2)
		self.over_todo_ratio = flt((over[4]), 2)

		gloss = self.calculate_attachments_details("gloss_level_details_", ['master_sample', 'finished_board_1', 'finished_board_2', 'finished_board_3', 'finished_board_4'], 
											 ['results_select_1', 'results_select_2', 'results_select_3', 'results_select_4'])
		self.gloss_total_attachment = gloss[0]
		self.gloss_pending_attachment = gloss[1]
		self.gloss_pass_ration = flt((gloss[2]), 2)
		self.gloss_undetermined_ratio = flt(gloss[3], 2)
		self.gloss_todo_ratio = flt((gloss[4]), 2)

		moisture = self.calculate_attachments_details("moisture_content_details_", 
												['master_sample', 'finished_board_1', 'finished_board_2', 'finished_board_3', 'finished_board_4'],
												['results_select_1', 'results_select_2', 'results_select_3', 'results_select_4'])
		self.moisture_total_attachment = moisture[0]
		self.moisture_pending_attachment = moisture[1]
		self.moisture_pass_ration = flt((moisture[2]), 2)
		self.moisture_undetermined_ratio = flt((moisture[3]), 2)
		self.moisture_todo_ratio = flt((moisture[4]), 2)

		open = self.calculate_attachments_details("open_box_inspection_details_", ['finished_board'],
											['bowing_select', 'squareness_select', 'ledging_overwood_select', 'pad_away_select'])
		self.open_total_attachment = open[0]
		self.open_pending_attachment = open[1]
		self.open_pass_ration = flt((open[2]), 2)
		self.open_undetermined_ratio = flt(open[3], 2)
		self.open_todo_ratio = flt((open[4]), 2)

		width = self.calculate_attachments_details("width_and_thickness_details_", 
											 ['finished_board_1', 'finished_board_2', 'finished_board_3', 'master_sample_matching_board'],
											 ['manual_select'])
		self.width_total_attachment = width[0]
		self.width_pending_attachment = width[1]
		self.width_pass_ration =flt(( width[2]), 2)
		self.width_undetermined_ratio = flt((width[3]), 2)
		self.width_todo_ratio = flt((width[4]), 2)

	def set_color_count_for_color_match_and_embossing(self):
		total_color = 0
		if self.get("no_of_po") and self.no_of_po > 0:
			for i in range(self.no_of_po):
				child_table_name="color_match_and_embossing_details_"+cstr(i+1)
				if len(self.get(child_table_name)) > 0:
					total_color = total_color + len(self.get(child_table_name))
		
		self.color_count = total_color

	@frappe.whitelist()
	def make_quality_control_item_using_tas_po_items(self, items):

		# empty child table values
		total_no_of_child_table=10
		for idx in range(total_no_of_child_table):
			child_table_name="quality_control_item_"+cstr(idx+1)
			parent_po_field_name="tas_po_name_"+cstr(idx+1)
			self.set(child_table_name,[])
			self.set(parent_po_field_name,None)

		# get tas po items
		if items:
			tas_po_list = frappe.db.get_list(
				"TAS Purchase Order Item",
				parent_doctype="TAS Purchase Order",
				filters={"name": ["in", items]},
				fields=["name", "parent", "item_no", "item_desc", "qty", "color", "cost"],
			)

			unique_tas_po = []
			for po in tas_po_list:
				if po.parent not in unique_tas_po:
					unique_tas_po.append(po.parent)

			self.no_of_po = len(unique_tas_po)

			# set tas po name
			for idx,unique in enumerate(unique_tas_po):
				parent_po_field_name="tas_po_name_"+cstr(idx+1)
				for po in tas_po_list:
					if unique == po.parent:
						self.set(parent_po_field_name,po.parent)
						break
			#  set items
			for idx,unique in enumerate(unique_tas_po):
				for po in tas_po_list:
					if unique == po.parent:
						child_table_name="quality_control_item_"+cstr(idx+1)
						po1 = self.append(child_table_name, {})
						po1.item = po.item_no
						po1.item_name = po.item_desc
						po1.qty = po.qty
						po1.color = po.color
						po1.amount = po.cost
						po1.item_color = cstr(po.item_no) + "-" +(cstr(po.color) or 'red')
						po1.tas_po_ref = po.parent

			self.set_pallet_details_table(unique_tas_po)

			self.set_po_ref_and_child_tables(unique_tas_po, "carton_po_", "inner_and_outer_carton_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_color_", "color_match_and_embossing_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_over_", "over_wax_and_edge_paint_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_gloss_", "gloss_level_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_moisture_", "moisture_content_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_open_", "open_box_inspection_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_width_", "width_and_thickness_details_")
			
			self.save(ignore_permissions=True)

		return self
	
	@frappe.whitelist()
	def make_quality_control_item_using_tas_po(self, tas_po):

		# empty child table values
		total_no_of_child_table=10
		for idx in range(total_no_of_child_table):
			child_table_name="quality_control_item_"+cstr(idx+1)
			parent_po_field_name="tas_po_name_"+cstr(idx+1)
			self.set(child_table_name,[])
			self.set(parent_po_field_name,None)

		if tas_po:
			tas_po_list = frappe.db.get_list(
				"TAS Purchase Order",
				filters={"name": ["in", tas_po]},
				fields=["name"],
			)

			unique_tas_po = []
			for po in tas_po_list:
				if po.name not in unique_tas_po:
					unique_tas_po.append(po.name)

			self.no_of_po = len(unique_tas_po)

			# set tas po name
			for idx,unique in enumerate(unique_tas_po):
				parent_po_field_name="tas_po_name_"+cstr(idx+1)
				for po in tas_po_list:
					if unique == po.name:
						self.set(parent_po_field_name,po.name)
						break
			#  set items
			for idx,unique in enumerate(unique_tas_po):
				for po in tas_po_list:
					if unique == po.name:
						child_table_name="quality_control_item_"+cstr(idx+1)
						po_doc = frappe.get_doc("TAS Purchase Order", po.name)

						for item in po_doc.items:
							po1 = self.append(child_table_name, {})
							po1.item = item.item_no
							po1.item_name = item.item_desc
							po1.qty = item.qty
							po1.color = item.color
							po1.amount = item.cost
							po1.item_color = cstr(item.item_no) + "-" + (cstr(item.color) or 'red')
							po1.tas_po_ref = item.parent
			
			self.set_pallet_details_table(unique_tas_po)

			self.set_po_ref_and_child_tables(unique_tas_po, "carton_po_", "inner_and_outer_carton_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_color_", "color_match_and_embossing_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_over_", "over_wax_and_edge_paint_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_gloss_", "gloss_level_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_moisture_", "moisture_content_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_open_", "open_box_inspection_details_")
			self.set_po_ref_and_child_tables(unique_tas_po, "po_width_", "width_and_thickness_details_")
			
			self.save(ignore_permissions=True)

		return self
	
	def set_po_ref_and_child_tables(self, po_list, parent_field, child_table):
		for idx,unique in enumerate(po_list):
			# parent po field
			parent_po_field_name=parent_field+cstr(idx+1)			
			self.set(parent_po_field_name,unique)

			# child table
			qi_items_table = "quality_control_item_"+cstr(idx+1)
			child_table_name=child_table+cstr(idx+1)

			# clear child table
			self.get(child_table_name).clear()

			# set child table
			if len(self.get(qi_items_table)) > 0:
				for item in self.get(qi_items_table):
					row = self.append(child_table_name, {})
					row.item_color = item.item_color

	def set_pallet_details_table(self, po_list):
		if len(po_list) > 0:
			self.pallet_details = []
			for po in po_list:
				row = self.append("pallet_details")
				row.tas_po = po


	def set_pallet_and_moisture_default_value(self):
		pallet_default_value = frappe.db.get_single_value('Quality Assurance Settings QI', 'pallet_default_value')
		open_box_guidelines = frappe.db.get_single_value('Quality Assurance Settings QI', 'open_box_inspection')

		if pallet_default_value:
			self.default_corner_width = ">=" + cstr(pallet_default_value)

		# moisture equipment default value
		default_value = ""
		if self.moisture_equipment and len(self.moisture_equipment) > 0:
			default_value = frappe.db.get_value('Equipment QI', self.moisture_equipment, 'equipment_default_value')
		
		self.default_moisture = default_value

		if open_box_guidelines:
			self.guidelines = open_box_guidelines
			


	def validate_over_wax_and_edge_paint_child_table(self):
		if self.get("no_of_po") and self.no_of_po > 0:
			for i in range(self.no_of_po):
				child_table_name="over_wax_and_edge_paint_"+cstr(i+1)
				if len(self.get(child_table_name)) > 0:
					for row in self.get(child_table_name):
						print(PASS_STATUS, '---PASS_STATUS')
						if row.over_wax_select not in PASS_STATUS and not row.finished_board:
							frappe.msgprint(_("In Over Wax Child Table, For {0} Item, Finished Board Picture is Require.").format(row.item_color), alert=True)



@frappe.whitelist()
def download_excel(doctype,docname,child_fieldname,file_name,data=None):
	ignore_fieldtype_in_list_view = ["Section Break", "Column Break", "HTML", "Button", "Image"]
	print(docname, '---docname')
	idx_no = 1
	qo_doc = frappe.get_doc(doctype, docname)
	file_header = ["Sr. NO."]
	file_data = []
	# file_data_list = [idx_no]
	for row in qo_doc.get(child_fieldname):
		file_data_list = [idx_no]
		child_doc = frappe.get_doc(row.doctype, row.name)
		print(child_doc, '---child_doc')
		print(qo_doc.meta.get_field(child_fieldname).label,"+++++++++++")
		field_data = child_doc.meta.fields
		for f in field_data:
			print(f.fieldtype, '---fieldtype')
			if not data:
				if f.in_list_view == 1:
					if f.fieldtype not in ignore_fieldtype_in_list_view:
						if f.label not in file_header:
							file_header.append(f.label)
						if f.fieldtype == "Check":
							print(child_doc.get(f.fieldname), '---fieldname')
							if child_doc.get(f.fieldname) == 1:
								print("Yes")
								file_data_list.append("Yes")
							else:
								print("No")
								file_data_list.append("No")
							
						elif f.fieldtype == "Attach":
							# file_header.append(f.label+" URL")
							file_data_list.append("""<img src={0} alt="img" width="500" height="600">""".format(child_doc.get(f.fieldname)))
						else:
							file_data_list.append(child_doc.get(f.fieldname))
			
			print(f.fieldname, '---fieldname',f.in_list_view)
		file_data.append(file_data_list)
		print(file_header, '---file_header')
		print(file_data, '---file_data')
		idx_no += 1

	public_file_path = frappe.get_site_path("public", "files")
	workbook = openpyxl.Workbook(write_only=True)
	# file_name=f"SBI-{docname}.xlsx"
	file_url=os.path.join(public_file_path,file_name)
	sheet = workbook.create_sheet(doctype, 0)
	sheet.append(file_header)
	for ele in file_data:
		sheet.append(ele)
	# sheet.append(file_footer)
	workbook.save(file_url)

	workBook = openpyxl.load_workbook(file_url)
	workSheet = workBook.active
	header_font_style = Font(bold=True, size=12, name="Calibri")
	color_code = "D3D3D3"
	# path = frappe.utils.get_url()+"/files/image.png"
	# img = openpyxl.drawing.image.Image(path) 
	# workSheet.add_image(img)
	for i in range(1, workSheet.max_column + 1):
		workSheet.cell(1, i).font = header_font_style
		workSheet.cell(1, i).fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")
		# workSheet.cell(workSheet.max_row, i).font = Font(bold=True, size=10, name="Calibri")
		workSheet.row_dimensions[1].height = 20
	
	for j in range(1, workSheet.max_row + 1):
		workSheet.cell(j, 1).font = header_font_style
		workSheet.cell(j, 1).fill = PatternFill(start_color=color_code, end_color=color_code, fill_type="solid")

	border_thin = Side(style='thin')
	for i in range (1, workSheet.max_row + 1):
		for j in range(1, workSheet.max_column + 1):
			workSheet.cell(i, j).alignment = Alignment(horizontal="center", vertical="center")
			workSheet.cell(i, j).border = Border(top=border_thin, left=border_thin, right=border_thin, bottom=border_thin)

	for column_cells in workSheet.columns:
		new_column_length = max(len(str(cell.value)) for cell in column_cells)
		new_column_letter = (chr(64+(column_cells[0].column)))
		if new_column_length > 0:
			workSheet.column_dimensions[new_column_letter].width = new_column_length+5 # *1.10

	workBook.save(file_url)



	return frappe.utils.get_url()+"/files/"+file_name
